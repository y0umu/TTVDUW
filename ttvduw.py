'''
This is The Very Document yoU Want (ttvduw)
'''
from pathlib import Path
import time
import csv
from copy import deepcopy
from docxtpl import DocxTemplate
from openpyxl import load_workbook

class DocuPrinter():
    '''
    读取 模板.docx，接受DataFeeder传入数据，输出格式化好的docx
    '''
    def __init__(self, tpl_name, out_path=None):
        '''
        tpl_name: 模板文件名。是一个docx文件
        out_path: 输出目录, str或None。如果不传入从str值则会将tpl_name的基名（PurePath.stem）作为输出目录的名称
        '''
        self.docu = DocxTemplate(tpl_name)
        self._ori_docx = deepcopy(self.docu.docx)   # hack。让同一个DocxTemplate()能多次使用
        self.context = {}  # 要填充的键值对，由self.set_context正确设置
        
        if out_path is None or len(out_path) == 0:
            out_path = Path(tpl_name).stem
        p_out_path = Path(out_path)
        p_out_path.mkdir(exist_ok=True)
        self.p_out_path = p_out_path

    
    def set_context(self, context: dict):
        '''
        context: 获取要填充的实际数据，每次获取的数据只填充一篇文档
        '''
        self.context.update(context)

    def write(self, keys=None):
        '''
        将渲染（填充）后的文件写入磁盘。写入的文件名从self.context中的键对应值选择（见keys参数）

        keys: list. 选择哪些键作为参数

        警告：这个方法不负责检查要输出的文件是否已经存在。如果已经存在，原文件将被覆盖。
        '''
        docu = self.docu
        docu.render(self.context)

        # 设置输出文件名
        out_name = ""
        use_fallback_filename = False
        # print(f'received keys: {keys}')
        if keys is None:
            use_fallback_filename = True
        else:
            for k in keys:
                _s = self.context.get(k)
                if _s is None:   # 用户输入了非法的键，跳过
                    continue
                out_name += str(_s)
                out_name += '_'
            if len(out_name) == 0: # 用户输入的全是非法键
                use_fallback_filename = True
        
        if use_fallback_filename:
            print('WARN: No valid keys given. Use fallback output filename')
            out_name = str(time.time() )   # fallback filename
        else:
            out_name = out_name[:-1]  # 去掉最后一个"_"字符
        
        out_name += '.docx'
        out_name = str(self.p_out_path / Path(out_name))
        docu.save(out_name)
        self.docu.docx = deepcopy(self._ori_docx)   # 重置DocxTemplate().docx使这个模板能再次使用


class DataFeeder():
    '''
    XlsxDataFeeder, CsvDataFeeder的父类
    '''
    def __init__(self, fname: str, 
                 tab_start_from_row=1,
                 tab_start_from_col=1):
        '''
        fname: str. 输入数据文件路径
        tab_start_from_row: int. 数据从第几行开始（默认：1）
        tab_start_from_col: int. 数据从第几列开始（默认：1）
        '''
        self.fname = fname
        self.min_row = tab_start_from_row
        self.min_col = tab_start_from_col
        self._keys = None      # 存储读取到的表格区键名，此类派生的子类负责设置此变量
    
    # def __enter__(self):
    #     '''
    #     如果子类需要，应该重新实现本方法
    #     '''
    #     return self
    
    # def __exit__(self, *exc_args):
    #     '''
    #     如果子类需要，应该重新实现本方法
    #     '''
    #     pass
    
    def _record_gen(self):
        '''
        子类应该重新实现本方法，使其返回一个可迭代对象，或者使本方法变成生成器。
        
        这个生成器应该生成list。每一个list都是键值数据表的一条记录。
        '''
        return []
    
    def load_file(self):
        '''
        子类应重新实现本方法， 使其装载datafeeder文件，并设置 self._keys的值
        '''
        self._keys = []

    def get_key(self, i: int):
        '''
        返回下标为i的self._keys中的元素
        '''
        return self._keys[i]
    
    def get_keys(self):
        '''
        返回self._keys
        '''
        return self._keys
    
    def context_feed(self, const_key_val={}):
        '''
        这是一个生成器，喂给DocxTemplate.set_context()
        @param const_key_val: dict。可选参数。用于补充常数键值映射。
        '''
        context = {}
        record_gen = self._record_gen()
        for r in record_gen:
            for k,v in zip(self._keys, r):
                context[k] = v
            if len(const_key_val) > 0:
                context.update(const_key_val)
            yield context



class XlsxDataFeeder(DataFeeder):
    '''
    读取数据xlsx表格的接口类
    '''
    def __init__(self, fname: str, 
                 tab_start_from_row=1, tab_start_from_col=1):
        '''
        fname: str. 输入数据文件路径
        tab_start_from_row: int. 数据从第几行开始（默认：1）
        tab_start_from_col: int. 数据从第几列开始（默认：1）
        '''
        super().__init__(fname, tab_start_from_row, tab_start_from_col)
        self._wb_xlsx = self.load_file()  # 存储xlsx的实例，load_file 创建之， __exit__ 销毁之

    def __enter__(self):
        return self
    
    def __exit__(self, *exc_args):
        '''
        exc_args: tuple of (exc_type, exc_val, exc_tb)
        '''
        self._wb_xlsx.close()
        print('debug: XlsxDataFeeder cleanned')

    def _record_gen(self):
        # keys_row, record_rows = self._get_ws_key_record_rows()

        # 将表格中的行转换成python list
        for r in self._record_rows:   # 余下的是每个记录具体的值
            ## 争议：空单元格应该直接返回None（默认行为）还是改写为""（空字符串）
            # 返回None时，如果模板中没有任何条件判断，就会印出"None"这四个字母
            # 返回""时，会让模板中对应位置什么有没有
            # this_row = [ x.value for x in r]
            this_row = [ x.value if x.value is not None else "" for x in r]

            yield this_row
    
    def _get_ws_key_record_rows(self):
        '''
        获取xlsx工作表中的键行，以及数据行生成器
        
        return:
        keys_row, ws_record_rows
        '''
        wb = self._wb_xlsx
        ws = wb.active
        # 部分软件生成的工作簿的 sheet1.xml 的 "dimension ref" 属性是错的
        # 修正参考如下链接
        # https://github.com/pandas-dev/pandas/issues/39001#issuecomment-762719332
        # https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/worksheet/worksheet.html#Worksheet.calculate_dimension
        # https://openpyxl.readthedocs.io/en/stable/optimized.html?highlight=reset_dimensions
        if ws.calculate_dimension() == "A1:A1":
            ws.reset_dimensions()
        ws_record_rows = ws.iter_rows(min_row=self.min_row, min_col=self.min_col)
        keys_row = next(ws_record_rows)  # 认为表格区的第1行是键名（字段名）
        return keys_row, ws_record_rows

    def load_file(self):
        '''
        装载 xlsx文件，并设置 self._keys
        '''
        wb = load_workbook(self.fname, read_only=True, data_only=True)
        self._wb_xlsx = wb
        keys_row, record_rows = self._get_ws_key_record_rows()
        self._record_rows = record_rows
        keys = [ str(x.value) for x in keys_row ]  # 转换成 python list
        self._keys = keys[(self.min_col-1):]
        
        return wb
    

class CsvDataFeeder(DataFeeder):
    '''
    读取数据xlsx表格的接口类
    '''
    def __init__(self, fname: str, 
                 tab_start_from_row=1, tab_start_from_col=1):
        '''
        fname: str. 输入数据文件路径
        tab_start_from_row: int. 数据从第几行开始（默认：1）
        tab_start_from_col: int. 数据从第几列开始（默认：1）
        '''
        super().__init__(fname, tab_start_from_row, tab_start_from_col)
        self._csvreader = None   # 由 load_file 设置
        self._csvfile = self.load_file()
        

    def __enter__(self):
        return self
    
    def __exit__(self, *exc_args):
        '''
        exc_args: tuple of (exc_type, exc_val, exc_tb)
        '''
        self._csvfile.close()
        print('debug: CsvDataFeeder cleanned')

    def load_file(self):
        csvfile = open(self.fname, newline='', encoding='utf-8')
        contents = csvfile.read(512)
        dialect = csv.Sniffer().sniff(contents)
        csvfile.seek(0)
        self._csvreader = csv.reader(csvfile, dialect)

        keys_row, self._record_rows = self._get_csv_key_record_rows()
        self._keys = keys_row[(self.min_col-1):]

        return csvfile

    def _get_csv_key_record_rows(self):
        '''
        获取csv中的键行，以及数据行生成器
        
        return:
        keys_row, record_rows_reader
        '''
        # 表格区不是从第1行开始的
        if self.min_row > 1:
            for _i in range(self.min_row - 1):
                next(self._csvreader)
        keys_row = next(self._csvreader) # 认为表格区的第1行是键名（字段名）
        record_rows_reader = self._csvreader
        return keys_row, record_rows_reader
    
    def _record_gen(self):
        for r in self._record_rows:   # 余下的是每个记录具体的值
            # print(f"debug: r == {r}")
            # 表格区不是从第1列开始的
            yield r[(self.min_col-1):]